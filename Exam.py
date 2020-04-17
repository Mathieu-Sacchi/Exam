'''
Exam correcion software
Py file and excel spreadsheet must be in same folder
'''


import openpyxl
wb = openpyxl.load_workbook(filename='Exam_results.xlsm',read_only=False, keep_vba=True)
ws = wb['Sheet1']
ws.append

'''
------------------------------------------------------------------------------
Question 1: Modify the function below so that it returns your last name.
If your last name has any non-alpha symbols (e.g. hyphen, spaces), remove them.
Example: The current function returns the value for the student named Jane Doe.
------------------------------------------------------------------------------
'''
class Student:

    def __init__(self,row):
        self.row = row
        self.ID = int(ws.cell(row = self.row + 1, column=4).value)
        self.L = ws.cell(row = self.row + 1, column=6).value

    def get_my_last_name(self):
        name = 'sacchi'  # must be lower-case
        #Used VBA to answer this question
        return name


    '''
    ------------------------------------------------------------------------------
    Question 2: Modify the function below so that it returns largest digit of
    your student id. You do not have to do any fancy computation; simply
    figure it out manually and hard-code the value.
    Example: The current function returns the value for the student named Jane Doe
    whose student ID is 45458212
    ------------------------------------------------------------------------------
    '''


    def get_my_largest_digit(self):

        ws.cell(row = self.row + 1, column=8).value = self.ID
        wb.save('Exam_results.xlsm')
        return ('test')



    '''
    ------------------------------------------------------------------------------
    Question 3: Modify the function below such that the parameter has a default
    value. That default value should be the number of alphabetic characters in
    your last name.
    Examples: For Jane Doe, num should have a default value of 3, thus
    get_my_successor() should return 4
    Other than that, the function should behave as it is:
    get_my_successor(0) should return 1
    get_my_successor(1) should return 2
    get_my_successor(-1) should return 0, and so on...
    ------------------------------------------------------------------------------
    '''


    def get_my_successor(self,num):
        successor = num + 1
        ws.cell(row=self.row + 1, column=9).value = successor
        wb.save('Exam_results.xlsm')
        # Used VBA to answer this question
        return successor

    '''
    ------------------------------------------------------------------------------
    Question 4: The area of a square with a width of w is (w x w). The area of
    n such squares is (n x w x w). Suppose that w is equal to the largest digit
    in your student ID. Modify the function below such that it returns the area of
    n squares with a width of w (w must be manually set within the function and
    should not be made into an extra parameter).Do not add any extra parameters to
    the function.
    Examples: For Jane, with student ID of 45458212 (largest digit is 8):
    get_area_of_my_squares(0) should return 0
    get_area_of_my_squares(1) should return 64
    get_area_of_my_squares(2) should return 128, and so on...
    ------------------------------------------------------------------------------
    '''


    def get_area_of_my_squares(self,n):

        area = self.ID*self.ID*n
        ws.cell(row=self.row + 1, column=10).value = area
        wb.save('Exam_results.xlsm')
        return area


    '''
    ------------------------------------------------------------------------------
    Question 5: The following function takes a string parameter as input and
    returns it back prefixed with a '('. Thus, calling it with 'hello' will return
    '(hello' Modify the function such that it prefixes and postfixes with your
    last name. Do not add any extra parameters to the function.
    Examples: For Jane, this is the expected behavior:
    get_my_modified_text('hello') should return 'doehellodoe'
    get_my_modified_text('EYE') should return 'doeEYEdoe'
    ------------------------------------------------------------------------------
    '''


    def get_my_modified_text(self,text):
        ws.cell(row=self.row + 1, column=11).value = ws.cell(row=self.row + 1, column=7).value + text + ws.cell(row=self.row + 1, column=7).value
        wb.save('Exam_results.xlsm')
        return


    '''
    ------------------------------------------------------------------------------
    Question 6: Modify the function such that it returns a string that repeats
    the first letter of your last name followed by the largest digit in your
    student ID, num times. Do not add any extra parameters to the function.
    Example: For Jane, this is the expected behavior:
    get_my_name_text(0) should return ''
    get_my_name_text(1) should return 'd8'
    get_my_name_text(5) should return 'd8d8d8d8d8'
    ------------------------------------------------------------------------------
    '''


    def get_my_name_text(self,num):
        txt=''
        for i in range (num):
            txt = txt + self.L + str(self.ID)

        ws.cell(row=self.row + 1, column=12).value = txt
        wb.save('Exam_results.xlsm')

        return txt


    '''
    ------------------------------------------------------------------------------
    Question 7: Modify the function such that:
        if num1 < N, it returns num2,
            where N is the largest digit in your student ID
        otherwise, it returns num3.
    
    Do not add any extra parameters to the function.
    Example: For Jane (N=8), this is the expected behavior:
    get_my_ternary(1,2,3) should return 2
    get_my_ternary(100,200,300) should return 300
    get_my_ternary(7,6,5) should return 6
    ------------------------------------------------------------------------------
    '''


    def get_my_ternary(self,num1, num2, num3):
        if num1 < self.ID:
            ws.cell(row=self.row + 1, column=13).value = num2
            wb.save('Exam_results.xlsm')
            return num2
        else:
            ws.cell(row=self.row + 1, column=13).value = num3
            wb.save('Exam_results.xlsm')
            return num3

    '''
    ------------------------------------------------------------------------------
    Question 8: Modify the function below such that it returns the average of all
    numbers between (N x a) and (N x b) (inclusive) where N is the largest digit
    in your student id. You may assume a < b. Do not add any extra parameters to
    the function.
    Example: For Jane, this is the expected behavior (N=8 for Jane):
    get_my_sequence_average(0,1) should return 4
    Note: result is (8x0 + 8x1)/2 = 4
    get_my_sequence_average(0,2) should return
    Note: result is (8x0 + 8x1 + 8x2)/3 = 8
    get_my_sequence_average(1,3) should return
    Note: result is (8x1 + 8x2 + 8x3)/3 = 16
    ------------------------------------------------------------------------------
    '''


    def get_my_sequence_average(self,a, b):
        sum = 0
        avrg = 0
        for i in range (a*self.ID,b*self.ID+1):
          sum = sum + i
        avrg = sum/(((b-a)*self.ID)+1)
        ws.cell(row=self.row + 1, column=14).value = avrg
        wb.save('Exam_results.xlsm')
        return avrg


    '''
    ------------------------------------------------------------------------------
    Question 9: Modify the function below such that it returns the count of numbers
    between a and b (exclusive) that are divisible by the ASCII value of the first
    letter of your last name (lowe-case). By exclusive it is implied that a and b
    will not be included. Do not add any extra parameters to the function.
    Example: For Jane, this is the expected behavior (ASCII value of d is 100):
    get_my_count(1,100) should return 0
    Note: because none of the numbers in sequence: 2,3,...,99 are divisible by 100
    get_my_count(1,150) should return 1
    Note: within the sequence: 2, 3, ..., 149, the number 100 is divisible by 100.
    get_my_count(1,1001) should return 10
    100, 200, 300, ..., 1000 are the 10 numbers present in the sequence: 2,...,1000
    ------------------------------------------------------------------------------
    '''


    def get_my_count(self,a, b):
        strg = str(self.L)
        ASCII = ord(str(strg))
        count = 0
        for i in range (a+1,b):
            if i % ASCII == 0:
                count += 1
        ws.cell(row=self.row + 1, column=15).value = count
        wb.save('Exam_results.xlsm')
        return count

    '''
    ------------------------------------------------------------------------------
    This recursive function is going to be used in Question 8.
    YOU ARE NOT ALLOWED TO MODIFY THIS FUNCTION.
    While working on Question 10, you will have to understand what this recursive
    function does.
    ------------------------------------------------------------------------------
    '''


    def repeat_text(self,text, count):
        if count <= 5:
            return ''
        else:
            return text + self.repeat_text(text, count - 1)


    '''
    ------------------------------------------------------------------------------
    Question 10: In the function below, Jane created a string where the first letter
    of her last name is repeated (5xN = 40) times, where N is the largest digit of
    her student ID. The function returns this string.
    Modify the function such that you return the first letter of your last name
    5xN times, where N is the largest digit of your student ID.
    
    There is a catch. You are not allowed to directly assign the letters the
    way Jane did. Jane cheated. She was supposed to call repeat_text with the
    appropriate parameters to generate her result. Figure out a way to call
    the recursive function repeat_text such that your letter is repeated 5xN times.
    Do not add any extra parameters to the function. Do not use any loops. And,
    don't cheat.
    '''


    def get_my_recursion(self):
        result = self.repeat_text(self.L, 5*(self.ID+1))
        # replace above with:
        # result = repeat_text(<something>,<something>)
        ws.cell(row=self.row + 1, column=16).value = result
        wb.save('Exam_results.xlsm')
        return result

#It seemed like dividing up the work helped at first but not really
for i in range(1,10):
    me = Student(i)
    me.get_my_largest_digit()
    me.get_my_successor(6)
    me.get_area_of_my_squares(4)
    me.get_my_modified_text("exam")
    me.get_my_name_text(3)
    me.get_my_ternary(1,2,3)
    me.get_my_sequence_average(1,20)
    me.get_my_count(1,110)
    me.get_my_recursion()
for i in range(10,20):
    me = Student(i)
    me.get_my_largest_digit()
    me.get_my_successor(6)
    me.get_area_of_my_squares(4)
    me.get_my_modified_text("exam")
    me.get_my_name_text(3)
    me.get_my_ternary(1,2,3)
    me.get_my_sequence_average(1,20)
    me.get_my_count(1,110)
    me.get_my_recursion()
for i in range(20,30):
    me = Student(i)
    me.get_my_largest_digit()
    me.get_my_successor(6)
    me.get_area_of_my_squares(4)
    me.get_my_modified_text("exam")
    me.get_my_name_text(3)
    me.get_my_ternary(1,2,3)
    me.get_my_sequence_average(1,20)
    me.get_my_count(1,110)
    me.get_my_recursion()



